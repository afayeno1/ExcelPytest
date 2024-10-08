#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os

import openpyxl
from openpyxl.reader.excel import load_workbook
import os

def read_case(dir):
    if os.path.isdir(dir):
        return read_folder(dir)
    else:
        return read_file(dir)

def read_folder(dir):
    data=[]
    for root,dirs,files in os.walk(dir):
        for file in files:
            if not file.startswith('.') and file!="element.xlsx":
                file_path=os.path.join(root,file)
                file_content=read_case(file_path)
                data.append(file_content)
    return data

def read_file(file):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook["Sheet1"]
    dict_data = {}
    for row in sheet.iter_rows():
        data = []
        for cell in row:
            if cell.column==1 and cell.value in ('模块','编号','用例名称','用例等级'):
                dict_data[cell.value]=cell.offset(column=1).value
                break
            elif cell.value=='步骤':
                continue
            else:
                if cell.column==2:
                    key = cell.value
                    dict_data[key] = []
                elif cell.value is not None:
                    if dict_data["模块"] in cell.value:
                        dict_data[key].append(get_xpath(cell.value))
                    else:
                        dict_data[key].append(cell.value)
                else:
                    break

    return dict_data

def element():
    wb = load_workbook("/Users/afayeno1/Documents/excelPytest/data_web/element.xlsx")
    element={}
    for sheet in wb.sheetnames:
        element[sheet] = {}
        for row in wb[sheet].iter_rows():
            for cell in row:
                element[sheet][cell.value] = cell.offset(column=1).value
                break
    return element

def get_xpath(data):
    xpath=element()[data.split(".")[0]][data.split(".")[1]]
    return xpath




# s=case_content("../data_web/登录/登录百度.xlsx")
# l=read_case("../data_web/登录/")
# print(l)
# s=element()
# print(s)
# x=get_xpath("登录.用户名")
# print(x)